from aiohttp import web
import asyncio
from db import get_all_users
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


async def index(request):
    users = await get_all_users()
    
    html = """
    <html>
    <head>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
            h1 { color: #333; }
            table { border-collapse: collapse; width: 100%; background: white; box-shadow: 0 0 10px rgba(0,0,0,0.1); }
            th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
            th { background: #4CAF50; color: white; }
            tr:nth-child(even) { background: #f2f2f2; }
            .export-btn { padding: 10px 20px; background: #4CAF50; color: white; text-decoration: none; border-radius: 5px; display: inline-block; margin: 20px 0; }
        </style>
    </head>
    <body>
        <h1>🏆 EBKA Championship - Участники</h1>
        <a class="export-btn" href="/export">📥 Экспорт в Excel</a>
        <table>
            <tr>
    """
    
    if users:
        keys = ["tg_id", "username", "epic", "discord", "rank", "peak_rank", "tracker"]
        headers = ["Telegram ID", "Username", "Epic ID", "Discord", "MMR", "Пик MMR", "Tracker"]
        for header in headers:
            html += f"<th>{header}</th>"
        html += "</tr>"
        
        for user in users:
            html += "<tr>"
            for key in keys:
                val = user.get(key, "—") or "—"
                html += f"<td>{val}</td>"
            html += "</tr>"
    else:
        html += "</tr><tr><td colspan='7'>Нет участников</td></tr>"
    
    html += """
        </table>
    </body>
    </html>
    """
    
    return web.Response(text=html, content_type='text/html')


async def export_excel(request):
    users = await get_all_users()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Участники"
    
    headers = ["Telegram ID", "Username", "Epic ID", "Discord", "MMR", "Пик MMR", "Tracker"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    keys = ["tg_id", "username", "epic", "discord", "rank", "peak_rank", "tracker"]
    for row_idx, user in enumerate(users, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=user.get(key, "—") or "—")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return web.Response(
        body=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=participants.xlsx'}
    )


async def handle_health(request):
    return web.Response(text="OK")


async def handle_participants(request):
    return await index(request)


async def handle_export(request):
    return await export_excel(request)


async def start_web():
    app = web.Application()
    app.router.add_route('GET', '/participants', handle_participants)
    app.router.add_route('GET', '/export', handle_export)
    app.router.add_route('GET', '/health', handle_health)
    
    runner = web.AppRunner(app)
    
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', 5000)
    await site.start()
    print("🌐 Веб-панель запущена на порту 5000")
    
    # Wait forever but handle cleanup properly
    stop_event = asyncio.Event()
    try:
        await stop_event.wait()
    except (asyncio.CancelledError, GeneratorExit, KeyboardInterrupt):
        print("🛑 Веб-сервер останавливается...")
    finally:
        try:
            await runner.cleanup()
            print("✅ Веб-сервер остановлен")
        except Exception as e:
            print(f"⚠️ Ошибка при остановке веб-сервера: {e}")